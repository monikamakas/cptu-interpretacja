from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import ScatterChart, Series, Reference
from openpyxl.chart.marker import Marker

CURVE_COLS = ["qc [MPa]", "fs [MPa]", "u2 [MPa]", "Rf(qc) [%]"]


def build_excel(df, meta=None):
    meta = meta or {}
    wb = Workbook()
    ws = wb.active
    ws.title = "Dane CPTU"

    ws["A1"] = "Wyniki sondowania statycznego CPTU"
    ws["A1"].font = Font(name="Arial", bold=True, size=13)
    row = 2
    for label, key in [("Numer testu", "test_number"), ("Nr stożka", "cone_number"),
                        ("Data", "date"), ("Inwestor", "investor"), ("Plik źródłowy", "source_file")]:
        val = meta.get(key)
        if val:
            ws[f"A{row}"] = f"{label}: {val}"
            ws[f"A{row}"].font = Font(name="Arial", size=9, italic=True, color="555555")
            row += 1
    ws[f"A{row}"] = "Dane odczytane wektorowo z wykresu PDF. Głębokość podana ze znakiem ujemnym (0 = powierzchnia terenu)."
    ws[f"A{row}"].font = Font(name="Arial", size=9, italic=True, color="555555")

    start_row = row + 2
    headers = list(df.columns)
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=start_row, column=j, value=h)
        c.font = Font(name="Arial", bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="2F5597")
        c.alignment = Alignment(horizontal="center")

    thin = Side(style="thin", color="D9D9D9")
    n = len(df)
    for i, r in df.iterrows():
        for j, h in enumerate(headers, start=1):
            val = r[h]
            cell = ws.cell(row=start_row + 1 + i, column=j,
                            value=round(float(val), 3) if val == val else None)
            cell.font = Font(name="Arial", size=10)
            cell.border = Border(bottom=thin)
            cell.number_format = "0.00" if j == 1 else "0.000"
            cell.alignment = Alignment(horizontal="center")

    widths = [16, 12, 12, 12, 13]
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = f"A{start_row + 1}"

    # --- Wykresy (jak na karcie PDF: wartość na X, głębokość na Y, głębokość ujemna) ---
    ws_chart = wb.create_sheet("Wykresy")
    depth_ref = Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + n)

    colors = {"qc [MPa]": "0000FF", "fs [MPa]": "FF0000", "u2 [MPa]": "800000", "Rf(qc) [%]": "000000"}
    for idx, col_name in enumerate(CURVE_COLS):
        col_idx = headers.index(col_name) + 1
        chart = ScatterChart()
        chart.title = col_name
        chart.style = 2
        chart.x_axis.title = col_name
        chart.y_axis.title = "Głębokość [m]"
        chart.height = 16
        chart.width = 8

        val_ref = Reference(ws, min_col=col_idx, min_row=start_row + 1, max_row=start_row + n)
        series = Series(depth_ref, val_ref, title=col_name)
        series.marker = Marker(symbol="none")
        series.graphicalProperties.line.width = 12000
        series.graphicalProperties.line.solidFill = colors.get(col_name, "000000")
        series.smooth = False
        # X = wartość parametru, Y = głębokość -> zamieniamy miejscami osie
        chart.series.append(series)
        chart.x_axis.crosses = "min"
        chart.y_axis.crosses = "min"
        anchor_col = get_column_letter(1 + idx * 6)
        ws_chart.add_chart(chart, f"{anchor_col}1")

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio
